import requests
from bs4 import BeautifulSoup
import openpyxl
import openpyxl.styles.numbers
from openpyxl.chart import BarChart, LineChart
import os
import xlsxwriter
from datetime import datetime
from time import gmtime, strftime
from collections import defaultdict
from openpyxl.chart import Reference
# import time if needed time.sleep()


HEADERS = {"user-agent": "GTA_SJ", "accept": '*/*'}
FILE = 'AutoRiaCars.xlsx'
HOST = 'https://auto.ria.com'
URL = 'https://auto.ria.com/uk/legkovie/chevrolet/camaro/'
cars = []
itemList = []


def get_html(url, params=None):
    r = requests.get(url, headers=HEADERS, params=params)
    return r


def get_pages_count(html):
    soup = BeautifulSoup(html, 'html.parser')
    pagination = soup.find_all('span', class_='mhide')
    if pagination:
        print(f'number of pages...:  {pagination[-1].get_text()} ')
        return int(pagination[-1].get_text())
    else:
        return 1


def get_content(html):
    soup = BeautifulSoup(html, 'html.parser')
    items = soup.find_all('div', class_='content-bar')

    cars = []

    for item in items:

        Mileage = item.find('li', class_='item-char js-race')
        Mileage = Mileage.get_text()[:-9]+'000'.replace(' ', '')
        if Mileage == ' без000':
            Mileage = ' 1000'

        # Test ADD_Time
        add_time = item.find('div', class_='footer_ticket')
        add_time = add_time.find_next('span').get('data-add-date')

        if add_time == None:
            add_time = strftime("%Y-%m-%d %H:%M:%S", gmtime())
        add_time = str(datetime.strptime(add_time, '%Y-%m-%d %H:%M:%S'))

        Year = item.find('a', class_='address')
        Year = Year.get_text()[-5:]

        cars.append({

            'Title': item.find('span', class_='blue bold').get_text(),
            'Link': item.find('a', class_='address').get('href'),
            'Price': item.find('div', class_='price-ticket').get('data-main-price'),  # + '$',
            'Year': Year,
            'Mileage': Mileage,
            'Engine': item.find('li', class_='item-char').find_next('li').find_next('li').get_text(),
            'City': item.find('li', class_='item-char view-location js-location').get_text()[:-8].replace(' ', ''),
            'Add_time': add_time,



        })
    print(cars)
    return cars


def dump_to_xlsx(filename, data, data_year_keys, data_year_coincidence):
    if not len(data):
        return None
    print(data_year_keys)
    print(data_year_coincidence)

    with xlsxwriter.Workbook(filename) as workbook:
        ws = workbook.add_worksheet()
        bold = workbook.add_format({'bold': True})

        '''for cell in ws['C']:
            cell.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]
        for cell in ws['D']:
            cell.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]
        for cell in ws['E']:
            cell.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]'''

        headers = ['Title', 'Year', ' Price $', 'Mileage (km)', 'Engine', 'Link', 'City', 'Add_time',
                   'Announcement time']

        for col, h in enumerate(headers):
            ws.write_string(0, col, h, cell_format=bold)

        for row, item in enumerate(data, start=1):
            ws.write_string(row, 0, item['Title'])
            ws.write_string(row, 5, item['Link'])
            ws.write_number(row, 2, int(item['Price']))
            ws.write_number(row, 3, int(item['Mileage']))
            ws.write_string(row, 4, item['Engine'])
            ws.write_number(row, 1, int(item['Year']))
            ws.write_string(row, 6, item['City'])
            ws.write_string(row, 7, item['Add_time'])
            ws.write_string(row, 8, str(datetime.strptime(item['Add_time'], '%Y-%m-%d %H:%M:%S') - datetime.now())[:-7])
            if row < len(data_year_keys)+1:
                ws.write_number(row, 10, int(data_year_keys[row-1]))
                ws.write_number(row, 11, int(data_year_coincidence[row - 1]))


def parse(URL):
    html = get_html(URL)
    if html.status_code == 200:

        pages_count = get_pages_count(html.text)
        for page in range(1, pages_count + 1):
            if page == 1:
                print(f'Now on the page: {page} of {pages_count}...')
                html = get_html(URL)
                cars.extend(get_content(html.text))

            else:
                print(f'Now on the page: {page} of {pages_count}...')
                html = get_html(URL, params={'page': page})
                cars.extend(get_content(html.text))

            # time.sleep(5) # pause for 5s
        for dictionary in cars:
            try:
                pass
                itemList.append(dictionary['Year'])
            except KeyError:
                pass

        # Find similar Years

        data_year = defaultdict(list)
        for i, item in enumerate(itemList):
            data_year[item].append(i)
        data_year = {k: v for k, v in data_year.items()}
        data_year = dict(sorted(data_year.items()))
        print(data_year)
        data_year_keys = []
        data_year_coincidence = []
        for i in data_year:
            data_year_keys.append(i)
            data_year_coincidence.append(len(data_year.get(i)))
        global max_len
        max_len = len(data_year_keys)
        try:
            dump_to_xlsx(FILE, cars, data_year_keys, data_year_coincidence)
        except KeyError:
            pass
        print(f'Get {len(cars)} cars')
        print(len(itemList))

    else:
        print('Error')


def create_figure():

    book = openpyxl.load_workbook("AutoRiaCars.xlsx")
    sheet = book.active

    chart = LineChart()

    data = Reference(sheet, min_col=3, max_col=4, min_row=2, max_row=len(cars)+1)
    chart.add_data(data)
    chart.y_axis.title = 'Mileage -- Price'
    chart.x_axis.title = 'Numbers Cars'

    sheet.add_chart(chart, 'J2')

    c1 = LineChart()
    c1.title = "Mileage -- Price"
    c1.y_axis.title = "Mileage"
    c1.x_axis.title = "Price"
    data1 = Reference(sheet, min_col=4, min_row=2, max_col=4, max_row=sheet.max_row)
    c1.add_data(data1, titles_from_data=True)
    dates = Reference(sheet, min_col=3, min_row=2, max_row=sheet.max_row)
    c1.set_categories(dates)
    sheet.add_chart(c1, "G17")
    c1.style = 13
    s1 = c1.series[0]
    s1.marker.symbol = "dot"
    s1.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
    s1.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline
    s1.graphicalProperties.line.noFill = True

    c2 = BarChart()
    c2.title = "Number of Number of matches"
    c2.y_axis.title = "Matches"
    c2.x_axis.title = "Year"
    data2 = Reference(sheet, min_col=12, min_row=1, max_col=12, max_row=max_len+1)
    c2.add_data(data2, titles_from_data=True)
    dates = Reference(sheet, min_col=11, min_row=2, max_row=max_len+1)
    c2.set_categories(dates)
    sheet.add_chart(c2, "J32")

    c3 = LineChart()
    c3.title = "Year -- Price"
    c3.y_axis.title = "Price"
    c3.x_axis.title = "Year"
    data3 = Reference(sheet, min_col=3, min_row=2, max_col=3, max_row=sheet.max_row)
    c3.add_data(data3, titles_from_data=True)
    dates = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row)
    c3.set_categories(dates)
    sheet.add_chart(c3, "J17")
    c3.style = 13
    s2 = c3.series[0]
    s2.marker.symbol = "dot"
    s2.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
    s2.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline
    s2.graphicalProperties.line.noFill = True

    book.save("AutoRiaCars.xlsx")


def main():
    parse(URL)
    create_figure()

    os.startfile("AutoRiaCars.xlsx")


if __name__ == '__main__':
    main()


