import requests
from bs4 import BeautifulSoup
import openpyxl
import openpyxl.styles.numbers
from openpyxl.chart import Reference, LineChart
import xlsxwriter
import os

URL = 'https://auto.ria.com/uk/newauto/marka-chevrolet/'
HEADERS = {"user-agent": 'ShprotBot'}
FILE = 'AutoRiaCars.xlsx'


def get_html(url, params=None):
    r = requests.get(url, headers=HEADERS, params=params)
    return r


def get_pages_count(html):
    soup = BeautifulSoup(html, "html.parser")
    pagination = soup.find_all('span', class_='mhide')
    if pagination:
        return int(pagination[-1].get_text())
    else:
        return 1


def get_content(html):
    soup = BeautifulSoup(html, "html.parser")
    items = soup.find_all('a', class_="proposition_link")
    cars = []

    for item in items:
        cars.append({
            'title': item.find('span', class_='link').get_text(strip=True),
            'description': item.find('div', class_='proposition_equip size13').get_text(),
            'year': item.find('span', class_='link').get_text(strip=True)[-4:],
            'price': item.find('span', class_='size16').get_text(strip=True).replace(" ", "")[:-3],
            'engine': item.find('span', class_='item').find_next('span').get_text(strip=True).replace('•', ':'),
            'region': item.find('span', class_='item region').get_text(strip=True),
        })

    return cars


def parse():
    html = get_html(URL)
    if html.status_code == 200:
        cars = []
        pages_count = get_pages_count(html.text)
        for page in range(1, pages_count+1):
            print(f" Szukam na stronie {page} з {pages_count}...")
            html = get_html(URL, params={'page': page})
            cars.extend(get_content(html.text))
        dump_to_xlsx(FILE, cars)
        print(f" Znalazłem tyle samochodów : {len(cars)}. ")
    else:
        print('Access Denied')


def dump_to_xlsx(filename, data):
    if not len(data):
        return None

    with xlsxwriter.Workbook(filename) as workbook:
        ws = workbook.add_worksheet()
        bold = workbook.add_format({'bold': True})

        headers = ['Title', 'Description', 'Year', 'Price', 'Engine', 'City']

        for col, h in enumerate(headers):
            ws.write_string(0, col, h, cell_format=bold)

        for row, item in enumerate(data, start=1):
            ws.write_string(row, 0, item['title'])
            ws.write_string(row, 1, item['description'])
            ws.write_number(row, 2, int(item['year']))
            ws.write_number(row, 3, int(item['price']))
            ws.write_string(row, 4, item['engine'])
            ws.write_string(row, 5, item['region'])


def create_figure():

    book = openpyxl.load_workbook("AutoRiaCars.xlsx")
    sheet = book.active

    c3 = LineChart()
    c3.title = "Year -- Price"
    c3.y_axis.title = "Price"
    c3.x_axis.title = "Year"
    data3 = Reference(sheet, min_col=4, min_row=2, max_col=4, max_row=sheet.max_row)
    c3.add_data(data3, titles_from_data=True)
    dates = Reference(sheet, min_col=3, min_row=2, max_row=sheet.max_row)
    c3.set_categories(dates)
    sheet.add_chart(c3, "J2")
    c3.style = 13
    s2 = c3.series[0]
    s2.marker.symbol = "dot"
    s2.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
    s2.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline
    s2.graphicalProperties.line.noFill = True

    book.save("AutoRiaCars.xlsx")


def main():
    parse()
    create_figure()
    os.startfile("AutoRiaCars.xlsx")


if __name__ == '__main__':
    main()
